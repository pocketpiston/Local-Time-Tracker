"""Generate a monthly Mehaffey Consulting invoice from time_tracker.db.

Usage:
    python generate_invoice.py                  # previous calendar month
    python generate_invoice.py --month last     # previous calendar month
    python generate_invoice.py --month 2026-05  # specific month
    python generate_invoice.py --month 2026-05 --out /tmp/foo.xlsx

Behavior:
  - Reads completed time logs from time_tracker.db whose [start,end) overlaps the month.
  - Cross-midnight sessions are split per calendar day; cross-month sessions are clipped
    so only the in-month portion is billed.
  - Groups by (date, project, item_code) and sums hours; descriptions joined with spaces.
  - Hours always round UP to the nearest 0.1h (6 min). No bucket ever rounds to zero.
  - Spillover < SPILLOVER_THRESHOLD_HOURS (0.15h) is folded back to the start day.
  - Item code is auto-classified from the description:
        "meeting"          -> Meeting
        "draft*|document*" -> Drafting (covers drafted/drafting/documented/documentation)
        otherwise          -> Research
  - Copies the OneDrive template and fills in the line items, invoice #, and date.
  - Hourly rate, GST, and totals are template formulas — they stay live.
  - Output saved to ./invoices/Mehaffey Invoice INV-YYYY-MM.xlsx unless --out is given.
  - Review the file before sending; tweak any Item/Description cells as needed.
"""

import argparse
import calendar
import datetime as dt
import math
import re
import shutil
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DB_PATH = HERE / "time_tracker.db"

TEMPLATE_PATH = Path(
    "/Users/joewu/Library/CloudStorage/OneDrive-Personal/Documents/"
    "[01] Current Documents/[03] Freelance Work/[01] Mehaffey Consulting/"
    "[01] Admin/[01] Billing/Mehaffy Billing Template.xlsx"
)

# OneDrive destination — a final copy lands here in the folder's own MMYYYY naming style.
# Set to None to skip the second copy.
ONEDRIVE_INVOICES_DIR = Path(
    "/Users/joewu/Library/CloudStorage/OneDrive-Personal/Documents/"
    "[01] Current Documents/[03] Freelance Work/[01] Mehaffey Consulting/"
    "[01] Admin/[01] Billing/[01] Invoices"
)

# Edit this list to skip projects you don't want billed (e.g. "Personal Projects").
EXCLUDE_PROJECTS: list[str] = []

OUTPUT_DIR = HERE / "invoices"

# Sessions that cross midnight are split between calendar days. If the post-midnight
# spillover is shorter than this threshold, fold those hours back into the start day
# instead of creating a tiny line item on the next day. 0.15h ~= 9 minutes.
SPILLOVER_THRESHOLD_HOURS = 0.15

FIRST_ITEM_ROW = 14
LAST_ITEM_ROW = 63  # template has 50 line-item rows (14..63)
INVOICE_NUM_CELL = "E3"
INVOICE_DATE_CELL = "E4"


def classify(description: str) -> str:
    d = (description or "").lower()
    if "meeting" in d:
        return "Meeting"
    if re.search(r"\b(draft\w*|document\w*)\b", d):
        return "Drafting"
    return "Research"


def month_bounds(year: int, month: int) -> tuple[dt.datetime, dt.datetime]:
    """Return (month_start, next_month_start) as naive datetimes — used for clipping."""
    start = dt.datetime(year, month, 1)
    if month == 12:
        next_start = dt.datetime(year + 1, 1, 1)
    else:
        next_start = dt.datetime(year, month + 1, 1)
    return start, next_start


def previous_month(today: dt.date) -> tuple[int, int]:
    first = today.replace(day=1)
    prev = first - dt.timedelta(days=1)
    return prev.year, prev.month


def parse_month_arg(value: str, today: dt.date) -> tuple[int, int]:
    if value is None or value == "last":
        return previous_month(today)
    m = re.fullmatch(r"(\d{4})-(\d{1,2})", value)
    if not m:
        raise SystemExit(f"--month must be YYYY-MM or 'last', got {value!r}")
    y, mo = int(m.group(1)), int(m.group(2))
    if not 1 <= mo <= 12:
        raise SystemExit(f"month out of range: {mo}")
    return y, mo


def fetch_entries(month_start: dt.datetime, next_month_start: dt.datetime) -> list[tuple[str, str, str, str]]:
    """Return rows (project, start_iso, end_iso, description) for any session whose
    [start, end) interval overlaps the month. Cross-boundary sessions are clipped
    in aggregate_by_day so only the in-month portion is billed.
    """
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT project_name, start_time, end_time, COALESCE(description, '')
            FROM time_logs
            WHERE end_time IS NOT NULL
              AND start_time < ?
              AND end_time > ?
            ORDER BY start_time
            """,
            (next_month_start.isoformat(), month_start.isoformat()),
        )
        return cur.fetchall()


def split_by_calendar_day(start: dt.datetime, end: dt.datetime):
    """Yield (date, hours) chunks so each calendar day gets only the time worked on it.

    A session 23:05 → 00:12 returns (start.date(), 0.91h), (start.date()+1, 0.21h).
    """
    cur = start
    while cur < end:
        next_midnight = dt.datetime.combine(cur.date() + dt.timedelta(days=1), dt.time())
        chunk_end = min(end, next_midnight)
        hours = (chunk_end - cur).total_seconds() / 3600.0
        if hours > 0:
            yield cur.date(), hours
        cur = chunk_end


def aggregate_by_day(rows, month_start: dt.datetime, next_month_start: dt.datetime) -> list[dict]:
    """Group by (date, project, item_code). Returns list sorted by date then project.

    Sessions that cross midnight are split so each calendar day gets only its share.
    Sessions are clipped to [month_start, next_month_start) so cross-month bleed-over
    is never billed on the wrong invoice.
    """
    buckets: dict[tuple, dict] = defaultdict(
        lambda: {"hours": 0.0, "descs": []}
    )
    for project, start_iso, end_iso, desc in rows:
        if project in EXCLUDE_PROJECTS:
            continue
        if desc.strip() == "[Paused]":
            continue
        start = dt.datetime.fromisoformat(start_iso)
        end = dt.datetime.fromisoformat(end_iso)
        # Clip to month bounds so cross-boundary sessions only contribute their
        # in-month portion to this invoice.
        start = max(start, month_start)
        end = min(end, next_month_start)
        if end <= start:
            continue
        item = classify(desc)
        clean = desc.strip()
        chunks = list(split_by_calendar_day(start, end))
        if not chunks:
            continue
        primary_day = chunks[0][0]
        for i, (day, hours) in enumerate(chunks):
            is_spillover = i > 0
            # Fold sub-threshold spillover back into the start-day bucket; skip its
            # description so the receiving day doesn't get cluttered with prior-day text.
            if is_spillover and hours < SPILLOVER_THRESHOLD_HOURS:
                buckets[(primary_day, project, item)]["hours"] += hours
                continue
            b = buckets[(day, project, item)]
            b["hours"] += hours
            if clean and clean not in b["descs"]:
                b["descs"].append(clean)

    out = []
    for (day, project, item), b in sorted(buckets.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2])):
        # Always round UP to nearest 0.1h (6 min). Bucket has >0 raw hours, so result is ≥0.1h.
        hours = math.ceil(b["hours"] * 10) / 10
        if hours <= 0:
            continue
        joined = " ".join(b["descs"])
        # Avoid redundant "LRTM. LRTM..." prefix when descriptions already reference the project.
        if not joined.lower().startswith(project.lower()):
            joined = f"{project}. {joined}"
        out.append({
            "date": day,
            "project": project,
            "item": item,
            "hours": hours,
            "description": joined,
        })
    return out


def render_invoice(entries: list[dict], year: int, month: int, out_path: Path) -> None:
    if not TEMPLATE_PATH.exists():
        raise SystemExit(f"Template not found: {TEMPLATE_PATH}")
    shutil.copy(TEMPLATE_PATH, out_path)

    wb = load_workbook(out_path)
    ws = wb["Invoice"]

    capacity = LAST_ITEM_ROW - FIRST_ITEM_ROW + 1
    if len(entries) > capacity:
        print(
            f"WARNING: {len(entries)} aggregated rows exceed template capacity "
            f"({capacity}). Extra rows will be dropped — extend the template or "
            f"tighten EXCLUDE_PROJECTS.",
            file=sys.stderr,
        )
        entries = entries[:capacity]

    # Clear any sample data left in line-item rows (A..D); column E formulas are preserved.
    for r in range(FIRST_ITEM_ROW, LAST_ITEM_ROW + 1):
        for col in ("A", "B", "C", "D"):
            ws[f"{col}{r}"] = None

    # Fill rows
    for i, e in enumerate(entries):
        r = FIRST_ITEM_ROW + i
        ws[f"A{r}"] = e["date"].strftime("%d/%m/%Y")
        ws[f"B{r}"] = e["item"]
        ws[f"C{r}"] = e["description"]
        ws[f"D{r}"] = e["hours"]
        # E{r} formula already in template: =IF(D{r}="","",D{r}*$E$11)

    # Invoice header
    ws[INVOICE_NUM_CELL] = f"INV-{year:04d}-{month:02d}"
    last_day = calendar.monthrange(year, month)[1]
    ws[INVOICE_DATE_CELL] = dt.datetime(year, month, last_day)

    wb.save(out_path)


def main() -> None:
    p = argparse.ArgumentParser(description="Generate Mehaffey monthly invoice from time tracker DB.")
    p.add_argument("--month", default="last", help="YYYY-MM or 'last' (default: previous month)")
    p.add_argument("--out", type=Path, default=None, help="Output xlsx path (default: ./invoices/Mehaffey Invoice INV-YYYY-MM.xlsx)")
    args = p.parse_args()

    today = dt.date.today()
    year, month = parse_month_arg(args.month, today)
    month_start, next_month_start = month_bounds(year, month)

    rows = fetch_entries(month_start, next_month_start)
    entries = aggregate_by_day(rows, month_start, next_month_start)

    if not entries:
        print(f"No billable entries found for {year}-{month:02d}.")
        return

    if args.out is None:
        OUTPUT_DIR.mkdir(exist_ok=True)
        out_path = OUTPUT_DIR / f"Mehaffey Invoice INV-{year:04d}-{month:02d}.xlsx"
    else:
        out_path = args.out

    render_invoice(entries, year, month, out_path)

    total_hours = sum(e["hours"] for e in entries)
    print(f"Wrote {out_path}")
    print(f"  Month: {year}-{month:02d}")
    print(f"  Line items: {len(entries)}")
    print(f"  Total hours: {total_hours:.2f}")

    # Mirror to the OneDrive Invoices folder using its MMYYYY naming convention.
    if ONEDRIVE_INVOICES_DIR is not None and args.out is None:
        if ONEDRIVE_INVOICES_DIR.exists():
            onedrive_name = f"Joe - Mehaffey Invoice {month:02d}{year:04d}.xlsx"
            onedrive_path = ONEDRIVE_INVOICES_DIR / onedrive_name
            shutil.copy(out_path, onedrive_path)
            print(f"  Mirrored to: {onedrive_path}")
        else:
            print(f"  (skip mirror — OneDrive folder not found: {ONEDRIVE_INVOICES_DIR})", file=sys.stderr)

    print("  Review the file in Excel before sending — adjust Item codes / descriptions as needed.")


if __name__ == "__main__":
    main()
